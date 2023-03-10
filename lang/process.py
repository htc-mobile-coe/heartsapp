import sys
import json
from jsonpath_ng import jsonpath, parse, Child
from openpyxl import load_workbook
import re
import csv

max_columns = 1000
operation = sys.argv[1]

with open("settings.json", "r") as f:
    settings = json.loads(f.read())
    print(settings)

def getFullPath(child):
    list = []
    if child != None and isinstance(child.left, Child):
        list.extend(getFullPath(child.left))
    else:
        list.append(str(child.left))
    list.append(str(child.right))
    return list

def toJsonObj(jsonObj, path, value):
    splits = path.split("|")
    if(len(splits)==1):
        jsonObj[path] = value
    else:
        if splits[0] not in jsonObj.keys():
            jsonObj[splits[0]]={}
        toJsonObj(jsonObj[splits[0]], "|".join(splits[1:]), value)

if operation == "fwd":
    file = sys.argv[2]
    print(file)
    if file not in settings.keys():
        print("File: "+file+" settings not defined in settings.json")
    else:
        with open(file, "r") as f:
            file_json = json.loads(f.read())
            print(file_json)
        flat_map = {}
        prefix = ""
        if not file == settings["default_file"]:
            prefix = file+";"
        for pattern in settings[file]:
            jsonpath_expr = parse(pattern)
            for match in jsonpath_expr.find(file_json):
                key = prefix+"|".join(getFullPath(match.full_path))
                flat_map[key] = match.value
        print(flat_map)
if operation == "diff":
    file = sys.argv[2]
    print(file)
    if file not in settings.keys():
        print("File: "+file+" settings not defined in settings.json")
    else:
        with open(file, "r") as f:
            file_json = json.loads(f.read())
            print(file_json)
        flat_map = {}
        prefix = ""
        if not file == settings["default_file"]:
            prefix = file+";"
        for pattern in settings[file]:
            jsonpath_expr = parse(pattern)
            for match in jsonpath_expr.find(file_json):
                key = prefix+"|".join(getFullPath(match.full_path))
                flat_map[key] = match.value
        print(flat_map)
        with open("previous/"+file, "r") as f:
            file_json = json.loads(f.read())
            print(file_json)
        flat_map_old = {}
        prefix = ""
        if not file == settings["default_file"]:
            prefix = file+";"
        for pattern in settings[file]:
            jsonpath_expr = parse(pattern)
            for match in jsonpath_expr.find(file_json):
                key = prefix+"|".join(getFullPath(match.full_path))
                flat_map_old[key] = match.value
        print(flat_map_old)

        diffMap = {}
        for key in flat_map.keys():
            if key not in flat_map_old.keys() or flat_map_old[key].strip() != flat_map[key].strip():
                diffMap [key] = flat_map[key]

        print(diffMap)

        with open('diff/diff.csv', 'w', newline='') as csvfile:
            diffwriter = csv.writer(csvfile, delimiter=',',
                                     quoting=csv.QUOTE_MINIMAL)
            diffwriter.writerow(["Key","Value"])
            for key in diffMap.keys():
                diffwriter.writerow([key, diffMap[key]])

if operation == "rev":
    file = sys.argv[2]
    print(file)
    flat_map_langs = {}
    wb = load_workbook(filename = file,data_only=True)
    sheetMaster = None
    for sheet in wb.worksheets:
        lang = "en"
        col_trx = 2
        if sheet.title == "Master":
            lang = "en"
            col_trx = 2
            sheetMaster = sheet
        elif sheet.title == "Instructions":
            continue
        else:
            langMatch=re.search(r"\(([A-Za-z0-9_\-]+)\)", sheet.title)
            if langMatch == None:
                print("Error## Sheet: "+sheet.title +" is not in correct format for language sheet title. Skipping it!!")
                continue
            lang = langMatch.group(langMatch.lastindex)
            col_trx = 3
        # print(sheet)
        flat_map_langs[lang]={}
        print("Processing Sheet: "+sheet.title)
        for i in range(2, max_columns):
            if (sheet.cell(row=i,column=1).value == None  or sheet.cell(row=i,column=1).value == "")\
                    and\
                (sheet.cell(row=i,column=2).value == None  or sheet.cell(row=i,column=2).value == "")\
                    :
                continue
            if sheet.cell(row=i,column=1).value == None  or sheet.cell(row=i,column=1).value == "":
                continue
            englishValue = sheetMaster.cell(row=i,column=2).value
            langValue = sheet.cell(row=i,column=col_trx).value
            if englishValue == None or langValue == None:
                continue

            if langValue != englishValue:
                # print("***")
                # print(englishValue)
                # print("****")
                varMatch = re.search(r"\{\{([A-Za-z0-9_\\-]+)\}\}", str(englishValue))
                if varMatch != None:
                    for reg in varMatch.regs:
                        if englishValue.count(englishValue[reg[0]:reg[1]]) != langValue.count(englishValue[reg[0]:reg[1]]):
                            print("Error## Invalid value Value doesn't contain all params key:"+sheet.cell(row=i,column=1).value+" eng:"+englishValue+" lang:"+langValue+ "Ignoring it ")
                            continue
            flat_map_langs[lang][sheet.cell(row=i,column=1).value] = langValue
    # print(flat_map_langs)
    fileLangJson = {}
    for lang in flat_map_langs:
        for key in flat_map_langs[lang]:
            fileKey = key.split(";")
            fileName = settings["default_file"]
            modifiedKey = key
            if len(fileKey) != 1:
                fileName = fileKey[0]
                modifiedKey = fileKey[1]

            if fileName not in fileLangJson.keys():
                fileLangJson[fileName]={}
            if lang not in fileLangJson[fileName].keys():
                fileLangJson[fileName][lang]={}
            toJsonObj( fileLangJson[fileName][lang], modifiedKey, flat_map_langs[lang][key] )

    for fileName in fileLangJson.keys():
        for lang in fileLangJson[fileName].keys():
            if fileName == "en.json":
                with open("out/"+lang+".json", 'w') as outfile:
                    json.dump(fileLangJson[fileName][lang], outfile, indent=4,ensure_ascii=False)



